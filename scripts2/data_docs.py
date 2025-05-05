# Imports
import os
import json
import streamlit as st
from textwrap import dedent
from agno.agent import Agent
from agno.models.google import Gemini
from agno.tools.file import FileTools
from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd


def convert_to_csv(file_path:str):
   """
    Use this tool to convert the excel file to CSV.

    * file_path: Path to the Excel file to be converted
    """
   # Load the file  
   df = pd.read_excel(file_path).head(10)

   # Convert to CSV
   st.write("Converting to CSV... :leftwards_arrow_with_hook:")
   return df.to_csv('temp.csv', index=False)


# Custom Tool to add comments to the header of an Excel file
def add_comments_to_header(file_path:str, data_dict:dict="data_dict.json"):
    """
    Use this tool to add the data dictionary as comments to the header of an Excel file and save the final file.

    The function takes a file path as argument and adds the {data_dict.json} as comments to each cell
    Start counting from column 0
    in the first row of the Excel file, using the following format:    
        * Column Number: <column_number>
        * Column Name: <column_name>
        * Data Type: <data_type>
        * Description: <description>

    Parameters
    ----------
    * file_path : str
        The path to the Excel file to be processed
    * data_dict : dict
        The data dictionary containing the column number, column name, data type, description, and number of null values

    Returns
    -------
    None
    """

    # Load the workbook
    wb = load_workbook(file_path)

    # Get the active worksheet
    ws = wb.active

    # Iterate over each column in the first row (header)
    for n, col in enumerate(ws.iter_cols(min_row=1, max_row=1)):
        for header_cell in col:
            header_cell.comment = Comment(dedent(f"""\
                              ColName: {data_dict[str(n)]['ColName']}, 
                              DataType: {data_dict[str(n)]['DataType']},
                              Description: {data_dict[str(n)]['Description']}\
    """),'AI Agent')

    # Save the workbook
    st.write("Wiriting Data Dictionary... :page_facing_up:")
    st.write("Saving File... :floppy_disk:")
    wb.save('final.xlsx')
    # wb.save('/app/documents/final.xlsx') # if using docker


# Create the agent
def create_agent(apy_key):
    agent = Agent(
        # model=Gemini(id="gemini-2.0-flash", api_key=os.environ.get("GEMINI_API_KEY")),
        model=Gemini(id="gemini-2.0-flash", api_key=apy_key),
        description= dedent("""\
                            You are an agent that reads the csv dataset presented to you and 
                            determine the following information:
                            - The data types of each column
                            - The description of each column                   
                            - The number of null values in each column

                            Using the tools provided, create a data dictionary in JSON format that includes the below information:
                            {<ColNumber>: {ColName: <ColName>, DataType: <DataType>, Description: <Description>}}
                            \
                            """),
        tools=[convert_to_csv,
               FileTools(read_files=True, save_files=True),
               add_comments_to_header],
        show_tool_calls=True
        )

    return agent


# Run the agent
if __name__ == "__main__":
    
    # Config page Streamlit
    st.set_page_config(layout="centered", 
                       page_title="Data Docs", 
                       page_icon=":paperclip:",
                       initial_sidebar_state="expanded")
    
    # Title
    st.title("Data Docs :paperclip:")
    st.subheader("Generate a data dictionary for your Excel file.")
    st.caption("1. Enter your Gemini API key and the path of the Excel file on the sidebar.")
    st.caption("2. Run the agent.")
    st.caption("3. The agent will generate a data dictionary and add it as comments to the header of the Excel file.")
    st.caption("ColName: <ColName> | DataType: <DataType> | Description: <Description>")

    with st.sidebar:
        # Enter your Gemini API key
        st.caption("Enter your Gemini API key and the path of the Excel file.")
        api_key = st.text_input("Enter your Gemini API key: ")
        input = st.text_input("Enter the path of the Excel file: ")

        agent_run = st.button("Run")

    # Create the agent
    if agent_run:
        agent = create_agent(api_key)
    
        # Start the script
        st.write("Running Agent... :runner:")

        # Run the agent    
        agent.print_response(dedent(f"""\
                                1. Read the dataset {input} and convert it to a temporary CSV using the 'convert_to_csv' tool.
                                2. Use the temp.csv as input to create the data dictionary for the columns in the dataset. 
                                3. Using the FileTools tool, save the data dictionary to a file named 'data_dict.json'.
                                4. Use the file 'data_dict.json' as input to the tool 'add_comments_to_header' and add the data dictionary as comments to the header of the dataset.
                                5. Save the excel file with comments to a file named 'final.xlsx'.
                                \
                                """),
                        markdown=True)
        
        # Print the data dictionary
        st.write("Generating Data Dictionary... :page_facing_up:")
        with open('data_dict.json', 'r') as f:
            data_dict = json.load(f)
            st.json(data_dict, expanded=False)

        # Remove temporary files
        st.write("Removing temporary files... :wastebasket:")
        os.remove('temp.csv')
        os.remove('data_dict.json')
    
    # If file exists, show success message
    if os.path.exists('final.xlsx'):
        st.success("Done! :white_check_mark:")

